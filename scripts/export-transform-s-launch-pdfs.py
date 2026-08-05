from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path(
    r"G:\EndoTech Singapore\Hugh\Part Number List -S-series\Updated Part Numbers reflecting CT and CTHT and CTS to ET changes"
    r"\Updated Transform S Part Numbers - CT CT-HT ET and MicroPath Extension - First Pass 2026-08-05.xlsx"
)
WORDMARK = ROOT / "src" / "assets" / "transform-s-wordmark.png"
OUTPUT_DIR = ROOT / "output" / "pdf" / "transform-s-launch"
PUBLIC_DIR = ROOT / "public" / "downloads"

NAVY = colors.HexColor("#193650")
TEAL = colors.HexColor("#008896")
TEAL_DARK = colors.HexColor("#0A6C76")
PALE_TEAL = colors.HexColor("#DDF2F3")
PALE_BLUE = colors.HexColor("#F3F7F9")
MID = colors.HexColor("#566B7A")
GRID = colors.HexColor("#CBD9DF")
WHITE = colors.white


def register_fonts() -> None:
    font_dir = Path(r"C:\Windows\Fonts")
    candidates = {
        "EndoTech": font_dir / "arial.ttf",
        "EndoTech-Bold": font_dir / "arialbd.ttf",
        "EndoTech-Italic": font_dir / "ariali.ttf",
    }
    for name, path in candidates.items():
        if path.exists():
            pdfmetrics.registerFont(TTFont(name, str(path)))
    if "EndoTech" not in pdfmetrics.getRegisteredFontNames():
        raise FileNotFoundError("Arial font files are required to build the product PDFs.")


FAMILIES = [
    {
        "sheet": "Transform S ET",
        "title": "Transform S\u2122 ET",
        "subtitle": "Confirmed New Zealand article numbers",
        "summary": "ET rotary shaping files in confirmed individual and assorted pack configurations.",
        "filename": "EndoTech-NZ-Transform-S-ET-Part-Numbers.pdf",
    },
    {
        "sheet": "Transform S PT",
        "title": "Transform S\u2122 PT",
        "subtitle": "Confirmed New Zealand article numbers",
        "summary": "Progressive taper shaping files and confirmed assorted pack configurations.",
        "filename": "EndoTech-NZ-Transform-S-PT-Part-Numbers.pdf",
    },
    {
        "sheet": "Transform S MicroPath",
        "title": "Transform S\u2122 Micro-Path",
        "subtitle": "Confirmed New Zealand article numbers",
        "summary": "Rotary glide path files, including the confirmed MB2 short glide path configuration.",
        "filename": "EndoTech-NZ-Transform-S-Micro-Path-Part-Numbers.pdf",
    },
    {
        "sheet": "C Plus",
        "title": "Transform S\u2122 C+ Files",
        "subtitle": "Confirmed New Zealand article numbers",
        "summary": "Confirmed ultra-stiff stainless-steel hand file configurations for calcified-canal negotiation.",
        "filename": "EndoTech-NZ-Transform-S-C-Plus-Part-Numbers.pdf",
    },
    {
        "sheet": "K Files",
        "title": "Transform S\u2122 K-Files",
        "subtitle": "Confirmed New Zealand article numbers",
        "summary": "Confirmed stainless-steel K-File individual and assorted pack configurations.",
        "filename": "EndoTech-NZ-Transform-S-K-Files-Part-Numbers.pdf",
    },
]


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def row_requires_confirmation(sheet: str, values: list[object]) -> tuple[bool, str]:
    joined = " | ".join(clean(value) for value in values).upper()
    sku = clean(values[4])
    size = clean(values[8])

    if "TO CONFIRM" in joined:
        return True, "One or more source-workbook fields are marked TO CONFIRM."
    if sheet == "Transform S MicroPath" and "NEW PRODUCT CODE AND UDI REQUIRED" in joined:
        return True, "New product code, UDI and final manufacturing specification require confirmation."
    if sheet == "Transform S ET" and sku in {"TSET-150421RF", "TSET-150425RF", "TSET-150429RF"}:
        return True, "New ET 15/.04 product-code, UDI and manufacturing reference require confirmation."
    if sheet == "Transform S ET" and size == "15/.04":
        return True, "New ET 15/.04 product-code, UDI and manufacturing reference require confirmation."
    return False, ""


def load_family_rows(workbook, sheet_name: str) -> tuple[list[dict[str, object]], int]:
    sheet = workbook[sheet_name]
    confirmed: list[dict[str, object]] = []
    held = 0

    for row in sheet.iter_rows(min_row=2, max_col=29, values_only=True):
        if not clean(row[4]):
            continue
        requires_confirmation, _ = row_requires_confirmation(sheet_name, list(row))
        if requires_confirmation:
            held += 1
            continue
        confirmed.append(
            {
                "sku": clean(row[4]),
                "file_type": clean(row[7]),
                "size": clean(row[8]),
                "length": clean(row[9]),
                "pack": clean(row[10]),
            }
        )
    return confirmed, held


def make_styles():
    styles = getSampleStyleSheet()
    return {
        "kicker": ParagraphStyle(
            "kicker",
            parent=styles["Normal"],
            fontName="EndoTech-Bold",
            fontSize=8.5,
            leading=10,
            textColor=TEAL,
            spaceAfter=4 * mm,
        ),
        "title": ParagraphStyle(
            "title",
            parent=styles["Title"],
            fontName="EndoTech-Bold",
            fontSize=26,
            leading=29,
            textColor=NAVY,
            alignment=TA_LEFT,
            spaceAfter=2 * mm,
        ),
        "subtitle": ParagraphStyle(
            "subtitle",
            parent=styles["Normal"],
            fontName="EndoTech",
            fontSize=11.5,
            leading=15,
            textColor=MID,
            spaceAfter=6 * mm,
        ),
        "section": ParagraphStyle(
            "section",
            parent=styles["Heading2"],
            fontName="EndoTech-Bold",
            fontSize=16,
            leading=19,
            textColor=NAVY,
            spaceBefore=2 * mm,
            spaceAfter=3 * mm,
        ),
        "body": ParagraphStyle(
            "body",
            parent=styles["BodyText"],
            fontName="EndoTech",
            fontSize=9.5,
            leading=13,
            textColor=NAVY,
        ),
        "body_bold": ParagraphStyle(
            "body_bold",
            parent=styles["BodyText"],
            fontName="EndoTech-Bold",
            fontSize=10,
            leading=13,
            textColor=NAVY,
        ),
        "small": ParagraphStyle(
            "small",
            parent=styles["BodyText"],
            fontName="EndoTech",
            fontSize=7.7,
            leading=10,
            textColor=MID,
        ),
        "table_header": ParagraphStyle(
            "table_header",
            parent=styles["Normal"],
            fontName="EndoTech-Bold",
            fontSize=7.5,
            leading=9,
            textColor=WHITE,
        ),
        "table_cell": ParagraphStyle(
            "table_cell",
            parent=styles["Normal"],
            fontName="EndoTech",
            fontSize=7.6,
            leading=9.4,
            textColor=NAVY,
        ),
        "table_sku": ParagraphStyle(
            "table_sku",
            parent=styles["Normal"],
            fontName="EndoTech-Bold",
            fontSize=7.6,
            leading=9.4,
            textColor=NAVY,
        ),
    }


def page_chrome(canvas, doc) -> None:
    width, height = A4
    canvas.saveState()
    if doc.page == 1:
        canvas.setFillColor(TEAL)
        canvas.rect(0, height - 13 * mm, 48 * mm, 13 * mm, stroke=0, fill=1)
        canvas.setFillColor(NAVY)
        canvas.rect(48 * mm, height - 13 * mm, width - 48 * mm, 13 * mm, stroke=0, fill=1)
    else:
        canvas.setFillColor(NAVY)
        canvas.rect(0, height - 4 * mm, width, 4 * mm, stroke=0, fill=1)

    canvas.setStrokeColor(GRID)
    canvas.setLineWidth(0.5)
    canvas.line(15 * mm, 14 * mm, width - 15 * mm, 14 * mm)
    canvas.setFillColor(MID)
    canvas.setFont("EndoTech", 7.2)
    canvas.drawString(15 * mm, 9.5 * mm, "EndoTech NZ | Transform S product catalogue | 5 August 2026")
    canvas.drawRightString(width - 15 * mm, 9.5 * mm, str(doc.page))
    canvas.restoreState()


def summary_card(config, count: int, held: int, styles) -> Table:
    availability = (
        f"<b>{count}</b> confirmed article number{'s' if count != 1 else ''} are shown for account-order requests. "
        f"<b>{held}</b> workbook row{'s are' if held != 1 else ' is'} held from customer selection pending confirmation."
    )
    rows = [
        [Paragraph("CATALOGUE STATUS", styles["table_header"]), ""],
        [Paragraph(config["summary"], styles["body_bold"]), Paragraph(availability, styles["body"])],
    ]
    table = Table(rows, colWidths=[66 * mm, 108 * mm], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), TEAL_DARK),
                ("BACKGROUND", (0, 1), (1, 1), PALE_TEAL),
                ("BOX", (0, 0), (1, 1), 0.6, colors.HexColor("#8BCED2")),
                ("INNERGRID", (0, 1), (1, 1), 0.45, colors.HexColor("#A9D9DC")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5 * mm),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5 * mm),
                ("TOPPADDING", (0, 0), (-1, -1), 3.2 * mm),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3.2 * mm),
            ]
        )
    )
    return table


def catalogue_table(rows: list[dict[str, object]], styles) -> Table:
    data = [
        [
            Paragraph("FILE TYPE", styles["table_header"]),
            Paragraph("SIZE / TAPER", styles["table_header"]),
            Paragraph("LENGTH", styles["table_header"]),
            Paragraph("PACK", styles["table_header"]),
            Paragraph("ARTICLE NUMBER", styles["table_header"]),
        ]
    ]
    for row in rows:
        data.append(
            [
                Paragraph(row["file_type"], styles["table_cell"]),
                Paragraph(row["size"], styles["table_cell"]),
                Paragraph(f"{row['length']} mm" if row["length"] else "-", styles["table_cell"]),
                Paragraph(row["pack"], styles["table_cell"]),
                Paragraph(row["sku"], styles["table_sku"]),
            ]
        )
    table = Table(
        data,
        colWidths=[50 * mm, 39 * mm, 22 * mm, 16 * mm, 47 * mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("GRID", (0, 0), (-1, -1), 0.45, GRID),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.5 * mm),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.5 * mm),
        ("TOPPADDING", (0, 0), (-1, -1), 2.2 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2 * mm),
    ]
    for index in range(1, len(data)):
        commands.append(("BACKGROUND", (0, index), (-1, index), WHITE if index % 2 else PALE_BLUE))
    table.setStyle(TableStyle(commands))
    return table


def build_pdf(config, rows, held, styles) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / config["filename"]
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=22 * mm,
        bottomMargin=20 * mm,
        title=f"{config['title']} part numbers",
        author="EndoTech NZ",
        subject="Confirmed New Zealand Transform S launch catalogue",
    )

    story = [
        Spacer(1, 2 * mm),
        Image(str(WORDMARK), width=61 * mm, height=61 * mm * 116 / 742),
        Spacer(1, 5 * mm),
        Paragraph("NEW ZEALAND LAUNCH CATALOGUE | 5 AUGUST 2026", styles["kicker"]),
        Paragraph(config["title"], styles["title"]),
        Paragraph(config["subtitle"], styles["subtitle"]),
        summary_card(config, len(rows), held, styles),
        Spacer(1, 7 * mm),
        Paragraph("Confirmed article numbers", styles["section"]),
        catalogue_table(rows, styles),
        Spacer(1, 5 * mm),
        KeepTogether(
            [
                Paragraph("Ordering and document status", styles["section"]),
                Paragraph(
                    "Use the article number when requesting an account order from EndoTech NZ at steveshepherdnz@gmail.com. "
                    "Availability must be confirmed at the time of order. "
                    "This document is a product catalogue summary only. It is not an IFU, technique guide, regulatory approval, or substitute for approved clinical instructions.",
                    styles["body"],
                ),
                Spacer(1, 2 * mm),
                Paragraph(
                    "Product-specific technique and IFU documents will be published separately after clinical and regulatory approval.",
                    styles["small"],
                ),
            ]
        ),
    ]
    doc.build(story, onFirstPage=page_chrome, onLaterPages=page_chrome)
    shutil.copy2(output_path, PUBLIC_DIR / config["filename"])
    return output_path


def main() -> None:
    register_fonts()
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)
    if not WORDMARK.exists():
        raise FileNotFoundError(WORDMARK)

    workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
    styles = make_styles()
    total_rows = 0
    total_held = 0
    for config in FAMILIES:
        rows, held = load_family_rows(workbook, config["sheet"])
        total_rows += len(rows)
        total_held += held
        output_path = build_pdf(config, rows, held, styles)
        print(f"{output_path.name}: confirmed={len(rows)} held={held}")

    if total_rows != 121 or total_held != 38:
        raise RuntimeError(
            f"Workbook reconciliation failed: expected 121 confirmed and 38 held; got {total_rows} confirmed and {total_held} held."
        )
    print(f"TOTAL: confirmed={total_rows} held={total_held}")


if __name__ == "__main__":
    main()
